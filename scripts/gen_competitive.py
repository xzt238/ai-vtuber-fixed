import logging
import openpyxl, os
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

wb = openpyxl.Workbook()

# Colors
GREEN = PatternFill('solid', fgColor='C6EFCE')
RED = PatternFill('solid', fgColor='FFC7CE')
YELLOW = PatternFill('solid', fgColor='FFEB9C')
GRAY = PatternFill('solid', fgColor='D9D9D9')
HEADER = PatternFill('solid', fgColor='2B579A')
BORDER = Border(left=Side('thin','D0D0D0'), right=Side('thin','D0D0D0'),
                top=Side('thin','D0D0D0'), bottom=Side('thin','D0D0D0'))

def fill_color(v):
    if v == '\u2705': return GREEN
    if v == '\u274c': return RED
    if v == '\u26a0\ufe0f': return YELLOW
    if v == '\u2014': return GRAY
    return None

# ===== Sheet 1: Matrix =====
ws = wb.active
ws.title = '\u529f\u80fd\u5bf9\u6bd4\u77e9\u9635'
headers = ['\u529f\u80fd\u7ef4\u5ea6', '\u529f\u80fd\u70b9', 'GuguGaga', 'VSeeFace', 'Warudo', 'ChatVRM', 'OpenClaw', 'VTube Studio', 'Animaze', 'VRoid Studio']
for i, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=i, value=h)
    c.font = Font(bold=True, size=10, color='FFFFFF')
    c.fill = HEADER
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

data = [
    ('3D\u6e32\u67d3', 'VRM 0.x \u6a21\u578b\u52a0\u8f7d', '\u2705', '\u2705', '\u2705', '\u2705', '\u2014', '\u2014', '\u2705', '\u2014'),
    ('3D\u6e32\u67d3', 'Live2D \u6a21\u578b\u652f\u6301', '\u2705', '\u2014', '\u2014', '\u2014', '\u2014', '\u2705', '\u2705', '\u2014'),
    ('3D\u6e32\u67d3', '\u591a\u6a21\u578b\u53d8\u4f53\u5207\u6362', '\u2705', '\u2014', '\u26a0\ufe0f', '\u2014', '\u2014', '\u2014', '\u2014', '\u2014'),
    ('3D\u6e32\u67d3', '\u9f20\u6807\u62d6\u62fd\u65cb\u8f6c', '\u2705', '\u2705', '\u2705', '\u2705', '\u2014', '\u2014', '\u2705', '\u2014'),
    ('3D\u6e32\u67d3', '\u6eda\u8f6e\u7f29\u653e', '\u2705', '\u2705', '\u2705', '\u2705', '\u2014', '\u2014', '\u2705', '\u2014'),
    ('3D\u6e32\u67d3', '\u5f39\u7c27\u9aa8\u9abc\u7269\u7406', '\u2705', '\u2705', '\u2705', '\u2705', '\u2014', '\u2705', '\u2705', '\u2014'),
    ('3D\u6e32\u67d3', '\u7a0b\u5e8f\u5316\u5f85\u673a\u52a8\u753b', '\u2705', '\u2014', '\u2705', '\u2014', '\u2014', '\u2014', '\u2014', '\u2014'),
    ('3D\u6e32\u67d3', 'BlendShape \u8868\u60c5', '\u26a0\ufe0f', '\u2705', '\u2705', '\u2705', '\u2014', '\u2014', '\u2705', '\u2014'),
    ('3D\u6e32\u67d3', '\u5b9e\u65f6\u53c2\u6570\u8c03\u8282\u9762\u677f', '\u2705', '\u2014', '\u2705', '\u2014', '\u2014', '\u2014', '\u2014', '\u2014'),
    ('\u8ffd\u8e2a\u4ea4\u4e92', 'Webcam \u9762\u90e8\u6355\u6349', '\u2014', '\u2705', '\u2705', '\u2014', '\u2014', '\u2705', '\u2705', '\u2014'),
    ('\u8ffd\u8e2a\u4ea4\u4e92', '\u624b\u90e8\u8ffd\u8e2a', '\u2014', '\u2705', '\u2705', '\u2014', '\u2014', '\u2014', '\u26a0\ufe0f', '\u2014'),
    ('\u8ffd\u8e2a\u4ea4\u4e92', '\u5168\u8eab\u52a8\u6355', '\u2014', '\u2014', '\u26a0\ufe0f', '\u2014', '\u2014', '\u2014', '\u26a0\ufe0f', '\u2014'),
    ('AI\u5bf9\u8bdd', 'LLM \u5927\u8bed\u8a00\u6a21\u578b', '\u2705', '\u2014', '\u2014', '\u2705', '\u2705', '\u2014', '\u2014', '\u2014'),
    ('AI\u5bf9\u8bdd', '\u591a Provider \u5207\u6362', '\u2705', '\u2014', '\u2014', '\u26a0\ufe0f', '\u2705', '\u2014', '\u2014', '\u2014'),
    ('AI\u5bf9\u8bdd', 'Function Calling', '\u2705', '\u2014', '\u2014', '\u2014', '\u2705', '\u2014', '\u2014', '\u2014'),
    ('AI\u5bf9\u8bdd', '\u591a\u8f6e\u5bf9\u8bdd\u5386\u53f2', '\u2705', '\u2014', '\u2014', '\u2705', '\u2705', '\u2014', '\u2014', '\u2014'),
    ('AI\u5bf9\u8bdd', '\u4e3b\u52a8\u8bf4\u8bdd', '\u2705', '\u2014', '\u2014', '\u2014', '\u2014', '\u2014', '\u2014', '\u2014'),
    ('\u8bed\u97f3\u80fd\u529b', 'TTS \u8bed\u97f3\u5408\u6210', '\u2705', '\u2014', '\u2014', '\u2705', '\u26a0\ufe0f', '\u2014', '\u26a0\ufe0f', '\u2014'),
    ('\u8bed\u97f3\u80fd\u529b', '\u58f0\u97f3\u514b\u9686 (GPT-SoVITS)', '\u2705', '\u2014', '\u2014', '\u2014', '\u2014', '\u2014', '\u2014', '\u2014'),
    ('\u8bed\u97f3\u80fd\u529b', 'STT \u8bed\u97f3\u8bc6\u522b', '\u2705', '\u2014', '\u2014', '\u2705', '\u2705', '\u2014', '\u2014', '\u2014'),
    ('\u8bed\u97f3\u80fd\u529b', '\u53e3\u578b\u540c\u6b65', '\u2705', '\u2705', '\u2705', '\u2705', '\u2014', '\u2705', '\u2705', '\u2014'),
    ('\u8bb0\u5fc6\u7cfb\u7edf', '\u77ed\u671f\u5de5\u4f5c\u8bb0\u5fc6', '\u2705', '\u2014', '\u2014', '\u2014', '\u2705', '\u2014', '\u2014', '\u2014'),
    ('\u8bb0\u5fc6\u7cfb\u7edf', '\u957f\u671f\u60c5\u666f\u8bb0\u5fc6', '\u2705', '\u2014', '\u2014', '\u2014', '\u2705', '\u2014', '\u2014', '\u2014'),
    ('\u8bb0\u5fc6\u7cfb\u7edf', '\u8bed\u4e49\u5411\u91cf\u68c0\u7d22', '\u2705', '\u2014', '\u2014', '\u2014', '\u2705', '\u2014', '\u2014', '\u2014'),
    ('\u8bb0\u5fc6\u7cfb\u7edf', '\u4e8b\u5b9e\u5e93\u63d0\u53d6', '\u2705', '\u2014', '\u2014', '\u2014', '\u2014', '\u2014', '\u2014', '\u2014'),
    ('\u8bb0\u5fc6\u7cfb\u7edf', '\u8bb0\u5fc6\u9057\u5fd8\u673a\u5236', '\u2705', '\u2014', '\u2014', '\u2014', '\u2014', '\u2014', '\u2014', '\u2014'),
    ('\u8bb0\u5fc6\u7cfb\u7edf', 'Dreaming/\u68a6\u5883\u6574\u5408', '\u2014', '\u2014', '\u2014', '\u2014', '\u2705', '\u2014', '\u2014', '\u2014'),
    ('\u8bb0\u5fc6\u7cfb\u7edf', '\u6bcf\u65e5\u53cd\u601d\u65e5\u8bb0', '\u2705', '\u2014', '\u2014', '\u2014', '\u26a0\ufe0f', '\u2014', '\u2014', '\u2014'),
    ('\u5e73\u53f0\u652f\u6301', '\u684c\u9762\u7aef', '\u2705', '\u2705', '\u2705', '\u2705', '\u2705', '\u2705', '\u2705', '\u2705'),
    ('\u5e73\u53f0\u652f\u6301', '\u6d4f\u89c8\u5668\u7aef', '\u2014', '\u2014', '\u2014', '\u2705', '\u2014', '\u2014', '\u2014', '\u2014'),
    ('\u5e73\u53f0\u652f\u6301', '\u591a\u5e73\u53f0\u6d88\u606f (Discord/Telegram)', '\u2014', '\u2014', '\u2014', '\u2014', '\u2705', '\u2014', '\u2014', '\u2014'),
    ('\u6d41\u5a92\u4f53', 'OBS \u865a\u62df\u6444\u50cf\u5934', '\u2014', '\u2705', '\u2705', '\u2014', '\u2014', '\u2014', '\u2705', '\u2014'),
    ('\u6d41\u5a92\u4f53', '\u76f4\u64ad\u5e73\u53f0\u4e92\u52a8', '\u2014', '\u2014', '\u2705', '\u2014', '\u26a0\ufe0f', '\u2014', '\u2014', '\u2014'),
    ('\u5176\u4ed6', '\u684c\u9762\u5ba0\u7269\u6a21\u5f0f', '\u2705', '\u2014', '\u2014', '\u2014', '\u2014', '\u2014', '\u2014', '\u2014'),
    ('\u5176\u4ed6', '\u6a21\u578b\u521b\u5efa/\u634f\u4eba', '\u2014', '\u2014', '\u2014', '\u2014', '\u2014', '\u2014', '\u2014', '\u2705'),
    ('\u5176\u4ed6', '\u5f00\u6e90', '\u26a0\ufe0f', '\u2705', '\u2014', '\u2705', '\u2705', '\u2014', '\u2014', '\u2705'),
    ('\u5176\u4ed6', '\u514d\u8d39', '\u2705', '\u2705', '\u26a0\ufe0f', '\u2705', '\u2705', '\u26a0\ufe0f', '\u26a0\ufe0f', '\u2705'),
]

for r, d in enumerate(data, 2):
    cat, feat, *vals = d
    for c, v in enumerate([cat, feat] + list(vals), 1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
        cell.font = Font(size=9)
        fc = fill_color(v)
        if fc: cell.fill = fc

for i, w in enumerate([12, 26, 14, 12, 12, 12, 12, 14, 12, 14], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A2'

# ===== Sheet 2: Gap =====
ws2 = wb.create_sheet('\u5dee\u8ddd\u5206\u6790')
for i, w in enumerate([40, 20, 20, 50], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

for i, h in enumerate(['\u5dee\u8ddd\u9879', '\u7f3a\u5931\u7a0b\u5ea6', '\u7ade\u54c1\u53c2\u8003', '\u5efa\u8bae'], 1):
    c = ws2.cell(row=1, column=i, value=h)
    c.font = Font(bold=True, size=11, color='FFFFFF')
    c.fill = HEADER
    c.alignment = Alignment(horizontal='center', vertical='center')

gaps = [
    ('\u9762\u90e8/\u8eab\u4f53\u8ffd\u8e2a', '\U0001f534 \u6838\u5fc3\u7f3a\u5931', 'VSeeFace / Warudo', '\u96c6\u6210 OpenSeeFace \u6216 MediaPipe \u5b9e\u73b0 webcam \u9762\u6355'),
    ('BlendShape \u8868\u60c5\u7cfb\u7edf', '\U0001f7e1 \u5f85\u5b8c\u5584', 'VSeeFace / VRoid Studio', '\u6362\u7528\u5e26 BlendShape \u7684 VRM \u6a21\u578b\uff0c\u5b9e\u73b0\u8868\u60c5\u5207\u6362'),
    ('OBS \u865a\u62df\u6444\u50cf\u5934\u8f93\u51fa', '\U0001f534 \u6838\u5fc3\u7f3a\u5931', 'VSeeFace / Animaze', '\u4f7f\u7528 pyvirtualcam \u6216 OBS VirtualCam SDK'),
    ('VMD/BVH \u52a8\u753b\u5bfc\u5165', '\U0001f7e1 \u5f85\u5b8c\u5584', 'VSeeFace / Warudo', '\u96c6\u6210 BVH \u89e3\u6790\u5668\u5230 three.js \u573a\u666f'),
    ('\u624b\u90e8\u8ffd\u8e2a (Leap/MediaPipe)', '\U0001f7e1 \u5f85\u5b8c\u5584', 'VSeeFace / Luppet', 'MediaPipe \u624b\u52bf\u8bc6\u522b \u2192 VRM \u624b\u81c2\u9aa8\u9abc'),
    ('\u591a\u5e73\u53f0\u6d88\u606f\u63a5\u5165', '\U0001f7e0 \u5dee\u8ddd\u8f83\u5927', 'OpenClaw', '\u96c6\u6210 Discord/Telegram Bot SDK'),
    ('Dreaming \u8bb0\u5fc6\u6574\u5408', '\U0001f7e1 \u5dee\u8ddd\u4e2d\u7b49', 'OpenClaw', '\u65e5\u8bb0 + \u4e3b\u52a8\u8bf4\u8bdd\u5df2\u6709\u57fa\u7840\uff0c\u589e\u52a0\u81ea\u52a8\u65f6\u6bb5\u603b\u7ed3'),
    ('\u6d41\u5a92\u4f53\u5e73\u53f0\u4e92\u52a8', '\U0001f7e0 \u5dee\u8ddd\u8f83\u5927', 'Warudo / Live3D', '\u63a5\u5165 Bilibili \u76f4\u64ad\u5f39\u5e55 \u2192 LLM \u2192 TTS \u2192 VRM'),
    ('\u6d4f\u89c8\u5668\u7aef/\u79fb\u52a8\u7aef', '\U0001f7e0 \u5dee\u8ddd\u8f83\u5927', 'ChatVRM / VTube Studio', 'WebSocket \u670d\u52a1\u5df2\u6709\u57fa\u7840\uff0c\u5c01\u88c5 PWA'),
    ('\u9762\u90e8\u8868\u60c5\u8bc6\u522b', '\U0001f7e1 \u5f85\u5b8c\u5584', 'VSeeFace', 'AnimationController \u60c5\u7eea\u4e8b\u4ef6\u5df2\u6709\uff0c\u5bf9\u63a5\u9762\u6355\u6570\u636e'),
]

for i, (item, sv, ref, sug) in enumerate(gaps, 2):
    ws2.cell(row=i, column=1, value=item).font = Font(size=10)
    ws2.cell(row=i, column=2, value=sv).font = Font(size=10, bold=True)
    ws2.cell(row=i, column=3, value=ref).font = Font(size=10, italic=True)
    ws2.cell(row=i, column=4, value=sug).font = Font(size=10)
    ws2.cell(row=i, column=4).alignment = Alignment(wrap_text=True)
    for j in range(1, 5):
        ws2.cell(row=i, column=j).border = BORDER
ws2.freeze_panes = 'A2'

# ===== Sheet 3: Strengths =====
ws3 = wb.create_sheet('\u4f18\u52bf\u5206\u6790')
ws3.column_dimensions['A'].width = 22
ws3.column_dimensions['B'].width = 65

for i, h in enumerate(['GuguGaga \u6838\u5fc3\u4f18\u52bf', '\u8be6\u60c5'], 1):
    c = ws3.cell(row=1, column=i, value=h)
    c.font = Font(bold=True, size=12, color='FFFFFF')
    c.fill = HEADER

strengths = [
    ('Live2D + VRM \u53cc\u6a21\u578b', '\u552f\u4e00\u540c\u65f6\u652f\u6301 Live2D Native OpenGL \u6e32\u67d3\u548c VRM 3D WebGL \u6e32\u67d3\u7684 AI VTuber \u9879\u76ee'),
    ('GPT-SoVITS \u58f0\u97f3\u514b\u9686', '\u7ade\u54c1\u4e2d\u552f\u4e00\u96c6\u6210\u672c\u5730\u58f0\u97f3\u514b\u9686\u5f15\u64ce\uff0c\u652f\u6301\u81ea\u5b9a\u4e49\u97f3\u8272'),
    ('\u56db\u5c42\u8bb0\u5fc6\u7cfb\u7edf', '\u5de5\u4f5c/\u60c5\u666f/\u8bed\u4e49/\u4e8b\u5b9e\u56db\u5c42\u67b6\u6784\uff0c\u5bf9\u6807 Mem0/Letta\uff0c\u7ade\u54c1\u4e2d\u65e0\u540c\u7b49\u65b9\u6848'),
    ('\u4e3b\u52a8\u8bf4\u8bdd', '\u57fa\u4e8e\u7a7a\u95f2\u8ba1\u65f6\u5668\u81ea\u52a8\u89e6\u53d1 LLM \u5bf9\u8bdd + TTS \u8bed\u97f3\u4ea4\u4e92\u95ed\u73af'),
    ('\u6bcf\u65e5\u53cd\u601d\u65e5\u8bb0', '\u7c7b\u4f3c OpenClaw Dreaming \u7684\u65e5\u53cd\u601d\u7cfb\u7edf\uff0cLLM \u81ea\u52a8\u603b\u7ed3\u5f53\u65e5\u4e8b\u4ef6'),
    ('Function Calling', 'AI \u53ef\u8c03\u7528\u672c\u5730\u5de5\u5177\uff08\u6587\u4ef6\u8bfb\u5199\u3001\u641c\u7d22\u3001\u547d\u4ee4\u6267\u884c\uff09'),
    ('VRM \u8bbe\u7f6e\u9875', '\u4e1a\u754c\u552f\u4e00\u7684 VRM \u6a21\u578b\u53c2\u6570\u5b9e\u65f6\u8c03\u8282\u9762\u677f\uff087\u9879\u53c2\u6570\uff09'),
    ('\u684c\u9762\u5ba0\u7269\u6a21\u5f0f', '\u53ef\u8131\u79bb\u7a97\u53e3\u4ee5\u684c\u9762\u5ba0\u7269\u5f62\u5f0f\u60ac\u6d6e\u8fd0\u884c'),
    ('\u591a LLM Provider', '\u652f\u6301 10+ \u63d0\u4f9b\u5546\uff08OpenAI/MiniMax/DeepSeek/Anthropic/Ollama\u7b49\uff09'),
]

for i, (t, d) in enumerate(strengths, 2):
    ws3.cell(row=i, column=1, value=t).font = Font(size=10, bold=True, color='2B579A')
    ws3.cell(row=i, column=2, value=d).font = Font(size=10)
    ws3.cell(row=i, column=2).alignment = Alignment(wrap_text=True)
    for j in range(1, 3):
        ws3.cell(row=i, column=j).border = BORDER

out = 'C:/Users/x/Desktop/ai-vtuber-fixed/docs/competitive_analysis.xlsx'
os.makedirs(os.path.dirname(out), exist_ok=True)
wb.save(out)
logger.info(f'Saved: {out} ({len(data)} features)')
